import os
import shutil
from time import sleep

import kivy
import pandas as pd

from pathlib import Path
from kivy.app import App
from kivy.uix.button import Button
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.spinner import Spinner
from kivy.uix.textinput import TextInput

from fetch_sgs_lib import start_fetch_period
from student_data_manager_lib import StudentData
from kivy.config import Config
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.scrollview import ScrollView
from kivy.core.window import Window
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.popup import Popup
import plyer
import asyncio
import threading


student_data = StudentData()
Config.set('graphics', 'default_font','THSarabunNew Bold')
Config.write()

class GPACalculator(GridLayout):
    def __init__(self):
        super(GPACalculator, self).__init__()
        self.rows = 1
        self.current_dataframe = None
        self.cols = 4
        self.gpa_label = Label(text="GPA")
        self.gpa_label.size_hint = (0.2,None)
        #gpax,steps = student_data.calculate_gpa(student_data.filter([student_data.possible_gpa["GPAX"]]))
        self.gpa_value_label = Label(text=str(0))
        self.gpa_value_label.size_hint = (0.2,None)

        self.select_gpa = SelectGPA()
        self.select_period = SelectPeriod()

        self.add_widget(self.gpa_label)
        self.add_widget(self.select_gpa)
        self.add_widget(self.select_period)
        self.add_widget(self.gpa_value_label)

        self.select_gpa.bind(text=self.update_label)
        self.select_period.bind(text = lambda x,t : self.update_label(None,self.select_gpa.text))
        self.update_label(None, "GPAX")

        global reupdate
        reupdate = lambda : self.reset_component()

    def update_label(self,spinner,text):
        if text not in student_data.possible_gpa:
            gpa, df = student_data.calculate_gpa(student_data.unordered_data)
            self.gpa_value_label.text = str(round(gpa, 2))
            self.current_dataframe = df
            self.on_update(df)
            return
        selected_filter = student_data.possible_gpa[text]
        period_filter = student_data.filter_by_period(int(self.select_period.text[0]))
        gpa,df = student_data.calculate_gpa(student_data.filter([*selected_filter,period_filter]))
        self.gpa_value_label.text = str("{:.2f}".format(gpa))
        self.current_dataframe = df
        self.on_update(df)

    def reset_component(self):
        self.clear_widgets()
        self.__init__()

    def on_update(self,df):
        pass

class SelectGPA(Spinner):
    def __init__(self):
        super(SelectGPA, self).__init__()
        self.text = "GPAX"
        self.size_hint = (0.2,None)

        self.font_name = "THSarabunNew Bold"
        self.option_cls.font_name = self.font_name
        self.font_size = 20
        self.option_cls.font_size = self.font_size

        values = []
        for field_name in student_data.possible_gpa:

            values.append(str(field_name))

        self.values = values

class SelectPeriod(Spinner):
    def __init__(self):
        super(SelectPeriod, self).__init__()
        self.text = f"{student_data.get_max_period()} ภาคเรียน"
        self.size_hint = (0.2,None)

        self.font_name = "THSarabunNew Bold"
        self.option_cls.font_name = self.font_name
        self.font_size = 20
        self.option_cls.font_size = self.font_size

        values = []
        for field_name in range(1,student_data.get_max_period()+1):
            values.append(f'{field_name} ภาคเรียน')
        self.values = values[::-1]

class ShowDF(BoxLayout):
    def graph(self):
        #df = pd.DataFrame.xls
        dfgui.show(self.df)

class GPALayout(BoxLayout):
    def __init__(self):
        super(GPALayout, self).__init__()
        self.orientation = "vertical"
        self.font_name = "THSarabunNew Bold"

        self.gpa_calculator = GPACalculator()
        #self.gpa_calculator.pos_hint = {"x":0,"y":0}
        self.add_widget(self.gpa_calculator)

        self.data_cell = ShowDataCell()
        self.data_cell.set_df(self.gpa_calculator.current_dataframe)
        self.gpa_calculator.on_update = self.data_cell.set_df
        #df_viewer.run()
        self.data_cell_holder = ScrollView(size_hint=(1, None), size=(Window.width, Window.height*0.7))
        #self.data_cell_holder.pos_hint = {"x":0,"y":0.2}
        self.data_cell_holder.add_widget(self.data_cell)
        self.add_widget(self.data_cell_holder)

        self.add_data_button = Button(text="เพิ่มข้อมูลภาคเรียน")
        self.save_as_excel_button = Button(text="Save As Excel")
        self.save_as_excel_button.size = (Window.width, Window.height*0.025)
        self.save_as_excel_button.bind(on_press=lambda *args:self.get_excel())
        self.add_data_button.font_name = "THSarabunNew Bold"
        #self.add_data_button.size = (Window.width, Window.height*0.05)
        self.add_data_obj = AddData()
        self.add_data_popup = Popup(content=self.add_data_obj)
        self.add_data_obj.exit_function = self.exit_function
        #self.add_data_popup.open()
        self.add_data_button.bind(on_press=lambda *args:self.add_data_popup.open())




        self.hold_add_data_and_excel_download = FloatLayout()
        self.hold_add_data_and_excel_download.add_widget(self.add_data_button)
        self.add_data_button.pos_hint = {"x":0,"y":0}
        self.add_data_button.size_hint_x = 0.8
        self.hold_add_data_and_excel_download.add_widget(self.save_as_excel_button)
        self.save_as_excel_button.pos_hint = {"y":0,"x":0.8}
        self.save_as_excel_button.size_hint_x = 0.2

        #self.add_widget(self.add_data_button)
        self.add_widget(self.hold_add_data_and_excel_download)

        #self.save_as_excel_button
        #self.add_widget(self.save_as_excel_button)

    def exit_function(self,*args):
        student_data.auto_read()
        student_data.auto_determined_possible()
        self.add_data_popup.dismiss()

    def get_excel(self):
        import openpyxl
        path = plyer.filechooser.save_file(path="a",filters=["xlsx"])[0]+".xlsx"
        self.gpa_calculator.current_dataframe.to_excel(path)

        wb = openpyxl.open(path)
        sheet = wb.worksheets[0]
        ptr_row = sheet.max_row+1
        sheet.cell(row=ptr_row,column=3).value = self.data_cell.sum_df[1]
        sheet.cell(row=ptr_row, column=4).value = self.data_cell.sum_df[2]

        wb.save(path)

class AddData(GridLayout):
    def __init__(self):
        super(AddData, self).__init__()
        self.cols = 1
        self.reset()

    def reset(self):
        self.clear_widgets()
        self.add_widget(SmartAddOrExit(self))
        for year in range(1,4):
            for period in range(1,3):
                self.first_period = SinglePeriod(year,period,f'{year}-{period}' in student_data.available_period,self)
                self.first_period.size_hint_y = 0.2
                self.add_widget(self.first_period)


class SmartAddOrExit(GridLayout):
    def __init__(self,add_data_obj):
        super(SmartAddOrExit, self).__init__()
        self.add_data_obj = add_data_obj
        self.cols = 2
        self.size_hint_y = 0.2
        self.start_fetch_button = Button(text="SmartDataFetch")
        self.start_fetch_button.bind(on_press=self.fetch_data)
        self.add_widget(self.start_fetch_button)
        self.exit_button = Button(text="Exit")
        self.exit_button.bind(on_press = lambda *args: self.add_data_obj.exit_function())
        self.add_widget(self.exit_button)
    
    def fetch_data(self,*args):
        fetch_data_form = FetchDataForm(self.add_data_obj)
        popup = Popup(content = fetch_data_form)
        popup.open()
        fetch_data_form.exit_function = popup.dismiss
    
class FetchDataForm(GridLayout):
    def __init__(self,add_data_obj):
        super(FetchDataForm, self).__init__()
        self.add_data_obj = add_data_obj
        self.cols = 2
        #self.rows = 2

        self.student_code_label = Label(text="รหัสนักเรียน")
        self.student_code_label.font_name = "THSarabunNew Bold"
        self.add_widget(self.student_code_label)

        self.student_code_field = TextInput()
        self.student_code_field.font_name = "THSarabunNew Bold"
        self.add_widget(self.student_code_field)

        self.citizen_code_label = Label(text="เลขประจําตัวประชาชน")
        self.citizen_code_label.font_name = "THSarabunNew Bold"
        self.add_widget(self.citizen_code_label)

        self.citizen_code_field = TextInput()
        self.citizen_code_field.password = True
        self.citizen_code_field.font_name = "THSarabunNew Bold"
        self.add_widget(self.citizen_code_field)

        self.exit_button = Button(text="Exit")
        self.exit_button.bind(on_press=lambda *args:self.exit_function())
        self.add_widget(self.exit_button)
        self.fetch_execute_button = Button(text = "Fetch")
        self.fetch_execute_button.bind(on_press = lambda *args : asyncio.run(self.fetch_period()))
        self.add_widget(self.fetch_execute_button)
        
    async def fetch_period(self):
        try:
            student_code,citizen_code = self.student_code_field.text, self.citizen_code_field.text
            #self.clear_widgets()
            progress_field = TextInput()
            #self.add_widget(progress_field)

            def progress_update(stage):
                progress_field.text += f"\n{stage}"

            start_fetch_period(student_code, citizen_code,progress_cb=progress_update)
        except BaseException as e:
            popup = Popup(
                title="Error",
                content=Label(text="รหัสนักเรียนหรือเลขประจําตัวประชาชนผิด", halign='left', valign='top'),
                size_hint=(None, None),
                size=(Window.width / 3, Window.height / 3),
                auto_dismiss=True,
            )
            popup.open()
            student_data.auto_read()
            student_data.auto_determined_possible()
            self.add_data_obj.reset()
            reupdate()
            return
        student_data.auto_read()
        student_data.auto_determined_possible()
        self.add_data_obj.reset()
        reupdate()
        self.exit_function()


class SinglePeriod(BoxLayout):
    def __init__(self,year,period,available,parent):
        super(SinglePeriod, self).__init__()
        self.add_data_obj = parent
        self.year = year
        self.period = period
        self.available = available
        self.size_hint = (1,0.2)

        period_name_text = f'[color=FF0000]ปี {self.year} ภาคเรียนที่ {self.period}[/color]' if not self.available else f'[color=00FF00]ปี {self.year} ภาคเรียนที่ {self.period}[/color]'

        self.period_name_label = Label(text=period_name_text,size=(Window.width,Window.height*0.2),markup=True)
        self.period_name_label.font_name = "THSarabunNew Bold"
        self.add_widget(self.period_name_label)

        self.set_this_period_button = Button(text="Set",size=(Window.width,Window.height*0.2))
        self.set_this_period_button.bind(on_press = self.add_data)
        self.set_this_period_button.font_name = "THSarabunNew Bold"

        self.del_this_period_button = Button(text="Delete", size=(Window.width, Window.height * 0.2))
        self.del_this_period_button.bind(on_press = self.delete_data)
        self.del_this_period_button.font_name = "THSarabunNew Bold"

        if not self.available:
            self.add_widget(self.set_this_period_button)
        else:
            self.add_widget(self.del_this_period_button)

    def add_data(self,*args):
        #print(__path__)
        path = plyer.filechooser.open_file()[0].split(os.sep)
        os.replace(os.path.join(*path), os.path.join(os.path.join(*os.path.abspath(__file__).split(os.sep)[:-1]),"data",f"{self.year}_{self.period}.xls"))

        #sleep(3)
        student_data.auto_read()
        self.add_data_obj.reset()

    def delete_data(self,*args):
        os.remove(f"data/{self.year}_{self.period}.xls")
        student_data.auto_read()
        self.add_data_obj.reset()


class TestApp(App):

    def build(self):
        self.title = 'FillTCASHelper'
        # return a Button() as a root widget
        return GPALayout()


class ShowDataCell(GridLayout):
    def __init__(self):
        super(ShowDataCell, self).__init__()
        self.cols = 7
        #self.size_hint = (1,None)
        self.bind(minimum_height=self.setter('height'))
        self.font_name = "THSarabunNew Bold"
        #self.height = self.minimum_height
        self.size_hint_y = None
        self.spacing = 10

    def add_cell(self,text):
        label = Label(text=str(text), size_hint_y=None)
        label.height = 48
        label.font_name = "THSarabunNew Bold"
        # self.size_hint_y = 0.2
        self.add_widget(label)

    def set_df(self,data):
        self.clear_widgets()
        data = data.dropna()
        sum_column = ["เกรด","หน่วยกิต","นน.xเกรด"]
        sum_df = data[sum_column].sum()
        all_columns = ["รายวิชา",*data.columns]
        for column in all_columns:
            self.add_cell(column)
        for subject_cell in data.iloc:
            self.add_cell(subject_cell.name)
            for data in subject_cell.to_numpy().tolist():
                self.add_cell(data)

        for column in all_columns:
            if column in sum_column:
                self.add_cell(sum_df[column])
            elif column == "รายวิชา":
                self.add_cell("ผลรวม")
            else:
                self.add_cell("")

        self.sum_df = sum_df









if __name__ == '__main__':
    student_data.auto_read()
    student_data.auto_determined_possible()
    TestApp().run()